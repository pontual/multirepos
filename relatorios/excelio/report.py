from datetime import date, timedelta, datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.writer.excel import save_virtual_workbook

from django.db import transaction
from django.db.models import Sum

from fichas.models import Produto, Incluido
from movimento.models import Chegando, Compra, Pedido, ItemPedido


def reportChegando(produto):
    produtoChegandos = Chegando.objects.filter(produto=produto)
    lenProdutoChegando = len(produtoChegandos)
    chegandoDisplay = ""
    if lenProdutoChegando > 0:
        for i, pc in enumerate(produtoChegandos):
            chegandoDisplay += "{} ({})".format(fmtThousands(pc.qtde), pc.nome)
            if i != lenProdutoChegando - 1:
                chegandoDisplay += ", "
    else:
        chegandoDisplay = ""
    return chegandoDisplay


def chegandoList(produto):
    produtoChegandos = Chegando.objects.filter(produto=produto)
    lenProdutoChegando = len(produtoChegandos)
    out = []
    if lenProdutoChegando > 0:
        for i, pc in enumerate(produtoChegandos):
            out.append("Chegando - {} ({})".format(fmtThousands(pc.qtde), pc.nome))
    return out


def totalChegando(produto):
    produtoChegandos = Chegando.objects.filter(produto=produto)
    total = 0
    for pc in produtoChegandos:
        total += pc.qtde
    return total


@transaction.atomic
def getBlocks(codigoBangs):
    """A bang! is an exclamation point used to mark codigos that
    should be checked manually"""
    
    # load everything into memory, there isn't that much data
    codigoBangs = list(map(str.upper, codigoBangs.split()))
    
    today = date.today()
    thisYear = today.year
    lastYear = thisYear - 1
    
    oneYearAgo = today - timedelta(days=365)
    lastYearBegin = date(lastYear, 1, 1)
    thisYearBegin = date(thisYear, 1, 1)
    startDate = lastYearBegin

    allVendas = ItemPedido.objects.filter(pedido__data__gte=startDate)
    vendas365 = allVendas.filter(pedido__data__gte=oneYearAgo)
    vendasLastYear = allVendas.filter(pedido__data__gte=lastYearBegin, pedido__data__lt=thisYearBegin)
    vendasThisYear = allVendas.filter(pedido__data__gte=thisYearBegin)

    response_err = ""
    thisyr = int(datetime.strftime(date.today(), "%Y"))
    lastyr = thisyr - 1

    blocks = []
    
    for codigo in codigoBangs:
        if not Produto.objects.filter(codigo=codigo).exists():
            response_err += "Warning: codigo {} not found\n".format(codigo)

    MONTH_AVG_FACTOR = 7

    for produto in Produto.objects.filter(inativo=False).order_by('codigo'):
        codigo = produto.codigo
        if codigo in codigoBangs:
            codigoDisplay = codigo + " (!)"
        else:
            codigoDisplay = codigo
            
        # vendasProdutoAll = allVendas.filter(produto__codigo=codigo)
        totalVendas365 = vendas365.filter(produto__codigo=codigo).aggregate(Sum('qtde')).get('qtde__sum')
        if totalVendas365 is None:
            totalVendas365 = 0
        
        totalVendasLastYear = vendasLastYear.filter(produto__codigo=codigo).aggregate(Sum('qtde')).get('qtde__sum')
        totalVendasThisYear = vendasThisYear.filter(produto__codigo=codigo).aggregate(Sum('qtde')).get('qtde__sum')

        if totalVendasLastYear is None:
            totalVendasLastYear = 0
            
        if totalVendasThisYear is None:
            totalVendasThisYear = 0

        # incluidos
        incluidos = Incluido.objects.filter(produto=produto).order_by('-data')[:3]
        incluido_str = ", ".join(datetime.strftime(i.data, "%d/%m/%y") for i in incluidos)
        
        # consider last "large" container (>= 9 boxes)
        cx2 = produto.cx * 2
        ultcont = Compra.objects.filter(produto=produto, qtde__gte=cx2).first()

        if ultcont:
            ultcontStr = "Ult cont: {} - {} - {}".format(
                fmtThousands(ultcont.qtde),
                datetime.strftime(ultcont.data, "%d/%m/%y"),
                ultcont.container)
        else:
            ultcontStr = ""

        firstcont = Compra.objects.filter(produto=produto).last()
        if ultcont is None:
            period_back_begin = oneYearAgo
            half = int(cx2 * 2)
        else:
            period_back_begin = max(oneYearAgo, firstcont.data)
            half = int(ultcont.qtde / 4)

        months_back = int(abs((today - period_back_begin).days) / 30)
        months_back = max(1, months_back)

        estoqueTotal = produto.disp + produto.resv
        # if produto.codigo in codigoBangs or vendas365 == 0 or estoqueTotal < half or (totalVendas365 / months_back * MONTH_AVG_FACTOR) > (estoqueTotal + totalChegando(produto)):

        includeCodigo = False
        
        # if produto.codigo in codigoBangs or estoqueTotal < half or (totalVendas365 / months_back * MONTH_AVG_FACTOR) > (estoqueTotal + totalChegando(produto)):

        if produto.codigo in codigoBangs:
            includeCodigo = True
        elif not produto.inativo and totalChegando(produto) == 0 and (estoqueTotal < half
                                                                      or (totalVendas365 / months_back * MONTH_AVG_FACTOR) > estoqueTotal):
            includeCodigo = True

        if includeCodigo:
            blocks.append({'codigo': codigo,
                           'codigodisp': codigoDisplay,
                           'nome': produto.nome.title(),
                           'chegando': reportChegando(produto),
                           'totalVendasLastYear': fmtThousands(totalVendasLastYear),
                           'totalVendasThisYear': fmtThousands(totalVendasThisYear),
                           'incluidos': incluido_str,
                           'ultimoest': fmtThousands(produto.ultimo_estoque),
                           'estoque': fmtThousands(estoqueTotal),
                           'disp': fmtThousands(produto.disp),
                           'resv': fmtThousands(produto.resv),
                           'ultcont': ultcontStr,
            })

        
    return blocks, response_err


@transaction.atomic
def semEstoque(produto):
    thisYear = date.today().year
    lastYear = thisYear - 1
    
    lastYearBegin = date(lastYear, 1, 1)

    changes = [(date.today(), 0)]

    saidas = ItemPedido.objects.filter(pedido__data__gte=lastYearBegin, produto=produto)
    entradas = Compra.objects.filter(data__gte=lastYearBegin, produto=produto)

    for s in saidas:
        changes.append((s.pedido.data, s.qtde))

    for e in entradas:
        changes.append((e.data, -e.qtde))

    changes.sort(reverse=True)

    # compute estoque
    curstock = produto.disp + produto.resv
    threshold = produto.cx // 2
    current_end_date = None

    date_ranges = []
    
    is_out = False
    for r in changes:
        curstock += r[1]
        if is_out:
            if curstock > threshold:
                is_out = False

                start_date_c = r[0]
                end_date_c = current_end_date

                if (end_date_c - start_date_c).days > 7:
                    if current_end_date == date.today():
                        display_date = "agora"
                    else:
                        display_date = datetime.strftime(current_end_date, "%d/%m/%y")
                    date_ranges.append("{} até {}".format(datetime.strftime(r[0], "%d/%m/%y"), display_date))
        if curstock < threshold and not is_out:
            is_out = True
            current_end_date = r[0]

    if date_ranges:
        return "Sem estoque de " + ", ".join(date_ranges)
    else:
        return ""
    

@transaction.atomic
def getXlsBlocks(cods):
    today = date.today()
    
    thisYear = today.year
    lastYear = thisYear - 1

    lastYearBegin = date(lastYear, 1, 1)
    thisYearBegin = date(thisYear, 1, 1)
    startDate = lastYearBegin

    allVendas = ItemPedido.objects.filter(pedido__data__gte=startDate)
    vendasLastYear = allVendas.filter(pedido__data__gte=lastYearBegin, pedido__data__lt=thisYearBegin)
    vendasThisYear = allVendas.filter(pedido__data__gte=thisYearBegin)

    thisyr = int(datetime.strftime(date.today(), "%Y"))
    lastyr = thisyr - 1

    blocks = []
    
    for produto in Produto.objects.filter(codigo__in=cods).order_by('codigo'):
        Incluido.objects.create(produto=produto, data=today)
        codigo = produto.codigo
        produto.ultimo_estoque = produto.disp + produto.resv
        produto.save()
        
        totalVendasLastYear = vendasLastYear.filter(produto__codigo=codigo).aggregate(Sum('qtde')).get('qtde__sum')
        totalVendasThisYear = vendasThisYear.filter(produto__codigo=codigo).aggregate(Sum('qtde')).get('qtde__sum')

        if totalVendasLastYear is None:
            totalVendasLastYear = 0
            
        if totalVendasThisYear is None:
            totalVendasThisYear = 0
            
        cx2 = produto.cx * 2
        ultcont = Compra.objects.filter(produto=produto, qtde__gte=cx2).first()

        if ultcont:
            ultcontStr = "Último container - {} - {} - {} pçs".format(
                datetime.strftime(ultcont.data, "%d/%m/%y"),
                ultcont.container,
                fmtThousands(ultcont.qtde))
        else:
            ultcontStr = ""
            
        threelargestsales = ItemPedido.objects.filter(produto=produto, pedido__data__gte=lastYearBegin).order_by('-qtde')[:3]
        threelargestsales = sorted(threelargestsales, key=lambda o: o.pedido.data, reverse=True)
        threelargest = ["{} - {} pçs - {}".format(datetime.strftime(s.pedido.data, "%d/%m/%y"), fmtThousands(s.qtde), s.pedido.cliente) for s in threelargestsales]

        semestoque = semEstoque(produto)
        
        blocks.append({'codigo': codigo,
                       'disp': produto.disp,
                       'resv': produto.resv,
                       'nome': produto.nome.title(),
                       'chegando': chegandoList(produto),
                       'chegandoTot': totalChegando(produto),
                       'totalVendasLastYear': totalVendasLastYear,
                       'totalVendasThisYear': totalVendasThisYear,
                       'caixa': produto.cx,
                       'ultcont': ultcontStr,
                       'threelargest': threelargest,
                       'semestoque': semestoque,
        })
        
    return blocks


def fmtThousands(x):
    return format(x, ',d').replace(',', '.')


def generateXlsReport(blocks):
    wb = Workbook()
    ws = wb.active

    thisyr = int(datetime.strftime(date.today(), "%Y"))
    lastyr = thisyr - 1

    todayStr = datetime.strftime(date.today(), "%d.%m.%Y")
    ws.title = todayStr

    ws.oddHeader.center.text = "Reposição " + todayStr
    ws.evenHeader.center.text = "Reposição " + todayStr
    ws.oddFooter.center.text = "Página &[Page] de &N"
    ws.evenFooter.center.text = "Página &[Page] de &N"
    
    arial_font = Font(name="Arial", size=10)
    bold_font = Font(name="Arial", bold=True, size=10)
    boldrows = set()

    # define title row
    ws['A1'] = "Código"
    ws['B1'] = "Disponível"
    ws['C1'] = "Reserva"
    ws['D1'] = "Chegando"
    ws['E1'] = "Vendas " + str(lastyr)
    ws['F1'] = "Vendas " + str(thisyr)
    ws['G1'] = "Caixa"

    ws['B1'].alignment = Alignment(horizontal='center')
    ws['C1'].alignment = Alignment(horizontal='center')
    ws['D1'].alignment = Alignment(horizontal='center')
    ws['E1'].alignment = Alignment(horizontal='center')
    ws['F1'].alignment = Alignment(horizontal='center')
    ws['G1'].alignment = Alignment(horizontal='center')

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 11
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 11
    ws.column_dimensions['E'].width = 13
    ws.column_dimensions['F'].width = 13
    ws.column_dimensions['G'].width = 8
    
    for cell in ws["1:1"]:
        cell.font = bold_font
        boldrows.add(1)
        
    # begin writing data
    row = 2

    for cod in blocks:
        ws['A' + str(row)] = str(cod['codigo'])
        ws['B' + str(row)] = fmtThousands(cod['disp'])
        ws['C' + str(row)] = fmtThousands(cod['resv'])
        ws['D' + str(row)] = fmtThousands(cod['chegandoTot'])
        ws['E' + str(row)] = fmtThousands(cod['totalVendasLastYear'])
        ws['F' + str(row)] = fmtThousands(cod['totalVendasThisYear'])
        ws['G' + str(row)] = fmtThousands(cod['caixa'])
        
        ws['B' + str(row)].alignment = Alignment(horizontal='center')
        ws['C' + str(row)].alignment = Alignment(horizontal='center')
        ws['D' + str(row)].alignment = Alignment(horizontal='center')
        ws['E' + str(row)].alignment = Alignment(horizontal='center')
        ws['F' + str(row)].alignment = Alignment(horizontal='center')
        ws['G' + str(row)].alignment = Alignment(horizontal='center')
    
        for cell in ws[str(row) + ":" + str(row)]:
            cell.font = bold_font
        boldrows.add(row)

        row += 1
        ws['A' + str(row)] = cod['nome']

        if cod['ultcont']:
            row += 1
            ws['A' + str(row)] = cod['ultcont']

        # Sem estoque
        if cod['semestoque']:
            row += 1
            ws['A' + str(row)] = cod['semestoque']

        for ch in cod['chegando']:
            row += 1
            ws['A' + str(row)] = ch
                       
        for large in cod['threelargest']:
            row += 1
            ws['A' + str(row)] = large

        row += 2

    for r in range(1, row):
        if r not in boldrows:
            for cell in ws[str(r)+":"+str(r)]:
                cell.font = arial_font

    ws.print_title_rows = "1:1"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.print_options.gridLines = True
    
    return save_virtual_workbook(wb)
